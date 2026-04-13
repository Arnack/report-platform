"""Unit tests for report generators."""
import pytest
import os
import tempfile
from openpyxl import load_workbook

from reports.sales_report import SalesReportGenerator
from reports.api_usage_report import APIUsageReportGenerator


@pytest.fixture
def temp_file():
    """Create a temporary file path for testing."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    yield path
    if os.path.exists(path):
        os.unlink(path)


@pytest.mark.asyncio
async def test_sales_report_generator(temp_file):
    """Test that SalesReportGenerator creates a valid XLSX file."""
    generator = SalesReportGenerator()
    
    # Verify properties
    assert generator.id == "sales_report"
    assert generator.name == "Sales Report"
    assert generator.description is not None
    assert generator.params_schema is not None
    
    # Generate report
    params = {"days": 30, "region": "All"}
    result_path = await generator.generate(params, temp_file)
    
    # Verify file was created
    assert os.path.exists(result_path)
    assert result_path.endswith('.xlsx')
    
    # Verify workbook structure
    wb = load_workbook(result_path)
    
    # Check sheets exist
    assert "Summary" in wb.sheetnames
    assert "Daily Trends" in wb.sheetnames
    
    # Check summary sheet has content
    ws = wb["Summary"]
    assert ws['A1'].value is not None  # Title should exist
    
    # Verify multiple rows of data
    data_rows = [row for row in ws.iter_rows(min_row=5, max_col=2, values_only=True) if row[0] is not None]
    assert len(data_rows) > 0


@pytest.mark.asyncio
async def test_sales_report_with_region_filter(temp_file):
    """Test SalesReportGenerator with region filter."""
    generator = SalesReportGenerator()
    params = {"days": 7, "region": "North"}
    
    result_path = await generator.generate(params, temp_file)
    assert os.path.exists(result_path)
    
    wb = load_workbook(result_path)
    assert "Summary" in wb.sheetnames


@pytest.mark.asyncio
async def test_api_usage_report_generator(temp_file):
    """Test that APIUsageReportGenerator creates a valid XLSX file."""
    generator = APIUsageReportGenerator()
    
    # Verify properties
    assert generator.id == "api_usage_report"
    assert generator.name == "API Usage Report"
    assert generator.description is not None
    
    # Generate report
    params = {"days": 7, "environment": "production"}
    result_path = await generator.generate(params, temp_file)
    
    # Verify file was created
    assert os.path.exists(result_path)
    
    # Verify workbook structure
    wb = load_workbook(result_path)
    
    # Check sheets exist
    assert "Overview" in wb.sheetnames
    assert "Daily Trends" in wb.sheetnames
    assert "Status Codes" in wb.sheetnames
    
    # Check overview sheet has content
    ws = wb["Overview"]
    assert ws['A1'].value is not None


@pytest.mark.asyncio
async def test_api_usage_report_default_params(temp_file):
    """Test APIUsageReportGenerator with default params."""
    generator = APIUsageReportGenerator()
    params = {}
    
    result_path = await generator.generate(params, temp_file)
    assert os.path.exists(result_path)
    
    wb = load_workbook(result_path)
    assert len(wb.sheetnames) >= 3


@pytest.mark.asyncio
async def test_sales_report_daily_trends(temp_file):
    """Test that SalesReport includes daily trends data."""
    generator = SalesReportGenerator()
    params = {"days": 14}
    
    result_path = await generator.generate(params, temp_file)
    wb = load_workbook(result_path)
    
    # Check Daily Trends sheet
    ws = wb["Daily Trends"]
    assert ws['A1'].value == "Date"
    assert ws['B1'].value == "Revenue"
    
    # Should have 14 rows of data (plus header)
    data_rows = [row for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0] is not None]
    assert len(data_rows) == 14


@pytest.mark.asyncio
async def test_api_usage_report_has_chart(temp_file):
    """Test that APIUsageReport includes charts."""
    generator = APIUsageReportGenerator()
    params = {"days": 7}
    
    result_path = await generator.generate(params, temp_file)
    wb = load_workbook(result_path)
    
    # Check Daily Trends sheet has chart
    ws = wb["Daily Trends"]
    assert len(ws._charts) > 0
