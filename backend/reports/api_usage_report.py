from typing import Dict, Any, Optional
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

from reports.base import BaseReportGenerator


class APIUsageReportGenerator(BaseReportGenerator):
    """API Usage report - generates XLSX with API usage analytics."""
    
    @property
    def id(self) -> str:
        return "api_usage_report"
    
    @property
    def name(self) -> str:
        return "API Usage Report"
    
    @property
    def description(self) -> str:
        return "API usage analytics with endpoint performance metrics and error rates"
    
    @property
    def params_schema(self) -> Optional[Dict[str, Any]]:
        return {
            "type": "object",
            "properties": {
                "days": {
                    "type": "integer",
                    "description": "Number of days to include in the report",
                    "default": 7,
                    "minimum": 1,
                    "maximum": 90
                },
                "environment": {
                    "type": "string",
                    "description": "Environment to report on",
                    "enum": ["production", "staging", "both"],
                    "default": "production"
                }
            }
        }
    
    async def generate(self, params: Dict[str, Any], output_path: str) -> str:
        days = params.get("days", 7)
        environment = params.get("environment", "production")
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14, color="4472C4")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        number_format = '#,##0'
        time_format = '0.00"ms"'
        
        # Overview Sheet
        ws_overview = wb.active
        ws_overview.title = "Overview"
        
        # Title
        ws_overview.merge_cells('A1:F1')
        title_cell = ws_overview['A1']
        title_cell.value = f"API Usage Report - {environment.upper()} - Last {days} days"
        title_cell.font = title_font
        ws_overview['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary metrics
        start_row = 4
        ws_overview[f'A{start_row}'] = "Summary Metrics"
        ws_overview[f'A{start_row}'].font = title_font
        
        total_requests = random.randint(50000, 500000)
        metrics = [
            ("Total Requests", total_requests, number_format),
            ("Successful Requests (2xx)", int(total_requests * 0.95), number_format),
            ("Client Errors (4xx)", int(total_requests * 0.03), number_format),
            ("Server Errors (5xx)", int(total_requests * 0.02), number_format),
            ("Average Response Time", round(random.uniform(50, 300), 2), time_format),
            ("P95 Response Time", round(random.uniform(200, 800), 2), time_format),
            ("P99 Response Time", round(random.uniform(500, 1500), 2), time_format),
        ]
        
        for i, (metric, value, fmt) in enumerate(metrics, start=start_row + 1):
            ws_overview[f'A{i}'] = metric
            ws_overview[f'B{i}'] = value
            ws_overview[f'B{i}'].number_format = fmt
        
        # Top Endpoints
        endpoint_row = start_row + 10
        ws_overview[f'A{endpoint_row}'] = "Top Endpoints by Usage"
        ws_overview[f'A{endpoint_row}'].font = title_font
        
        headers = ["Endpoint", "Method", "Requests", "Avg Response (ms)", "Error Rate (%)"]
        header_start = endpoint_row + 1
        for col, header in enumerate(headers, 1):
            cell = ws_overview.cell(row=header_start, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        endpoints = [
            ("/api/v1/users", "GET"),
            ("/api/v1/users", "POST"),
            ("/api/v1/reports", "GET"),
            ("/api/v1/reports", "POST"),
            ("/api/v1/analytics", "GET"),
            ("/api/v1/auth/login", "POST"),
            ("/api/v1/auth/refresh", "POST"),
            ("/api/v1/data/export", "POST"),
        ]
        
        for i, (endpoint, method) in enumerate(endpoints):
            row = header_start + 1 + i
            requests = random.randint(1000, 50000)
            avg_response = round(random.uniform(30, 500), 2)
            error_rate = round(random.uniform(0.1, 5.0), 2)
            
            ws_overview.cell(row=row, column=1, value=endpoint)
            ws_overview.cell(row=row, column=2, value=method)
            ws_overview.cell(row=row, column=3, value=requests).number_format = number_format
            ws_overview.cell(row=row, column=4, value=avg_response).number_format = time_format
            ws_overview.cell(row=row, column=5, value=error_rate).number_format = '0.00"%"'
        
        ws_overview.column_dimensions['A'].width = 30
        ws_overview.column_dimensions['B'].width = 12
        ws_overview.column_dimensions['C'].width = 15
        ws_overview.column_dimensions['D'].width = 22
        ws_overview.column_dimensions['E'].width = 18
        
        # Daily Trends Sheet
        ws_trends = wb.create_sheet("Daily Trends")
        ws_trends['A1'] = "Date"
        ws_trends['B1'] = "Requests"
        ws_trends['C1'] = "Avg Response (ms)"
        ws_trends['D1'] = "Errors"
        ws_trends['E1'] = "Error Rate (%)"
        
        for col in range(1, 6):
            cell = ws_trends.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        base_date = datetime.now() - timedelta(days=days)
        for i in range(days):
            row = 2 + i
            date = base_date + timedelta(days=i)
            requests = random.randint(5000, 50000)
            avg_response = round(random.uniform(50, 400), 2)
            errors = int(requests * random.uniform(0.01, 0.05))
            error_rate = round((errors / requests) * 100, 2)
            
            ws_trends.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
            ws_trends.cell(row=row, column=2, value=requests).number_format = number_format
            ws_trends.cell(row=row, column=3, value=avg_response).number_format = time_format
            ws_trends.cell(row=row, column=4, value=errors).number_format = number_format
            ws_trends.cell(row=row, column=5, value=error_rate).number_format = '0.00"%"'
        
        ws_trends.column_dimensions['A'].width = 15
        ws_trends.column_dimensions['B'].width = 15
        ws_trends.column_dimensions['C'].width = 22
        ws_trends.column_dimensions['D'].width = 12
        ws_trends.column_dimensions['E'].width = 15
        
        # Add chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Daily API Requests"
        chart.y_axis.title = "Requests"
        chart.x_axis.title = "Date"
        
        data = Reference(ws_trends, min_col=2, min_row=1, max_row=days + 1)
        cats = Reference(ws_trends, min_col=1, min_row=2, max_row=days + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_trends.add_chart(chart, "G2")
        
        # Status Codes Sheet
        ws_status = wb.create_sheet("Status Codes")
        ws_status['A1'] = "Status Code Distribution"
        ws_status['A1'].font = title_font
        
        headers = ["Status Code", "Description", "Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws_status.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        status_codes = [
            (200, "OK", 0.75),
            (201, "Created", 0.15),
            (204, "No Content", 0.05),
            (400, "Bad Request", 0.02),
            (401, "Unauthorized", 0.01),
            (403, "Forbidden", 0.005),
            (404, "Not Found", 0.01),
            (500, "Internal Server Error", 0.005),
        ]
        
        for i, (code, desc, pct) in enumerate(status_codes):
            row = 4 + i
            count = int(total_requests * pct)
            
            ws_status.cell(row=row, column=1, value=code)
            ws_status.cell(row=row, column=2, value=desc)
            ws_status.cell(row=row, column=3, value=count).number_format = number_format
            ws_status.cell(row=row, column=4, value=round(pct * 100, 2)).number_format = '0.00"%"'
        
        ws_status.column_dimensions['A'].width = 15
        ws_status.column_dimensions['B'].width = 25
        ws_status.column_dimensions['C'].width = 15
        ws_status.column_dimensions['D'].width = 15
        
        wb.save(output_path)
        return output_path
