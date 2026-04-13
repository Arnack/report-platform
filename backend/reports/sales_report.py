from typing import Dict, Any, Optional
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reports.base import BaseReportGenerator


class SalesReportGenerator(BaseReportGenerator):
    """Sales report - generates XLSX with sales data for a given period."""
    
    @property
    def id(self) -> str:
        return "sales_report"
    
    @property
    def name(self) -> str:
        return "Sales Report"
    
    @property
    def description(self) -> str:
        return "Comprehensive sales report with revenue breakdown by product and region"
    
    @property
    def params_schema(self) -> Optional[Dict[str, Any]]:
        return {
            "type": "object",
            "properties": {
                "days": {
                    "type": "integer",
                    "description": "Number of days to include in the report",
                    "default": 30,
                    "minimum": 1,
                    "maximum": 365
                },
                "region": {
                    "type": "string",
                    "description": "Filter by region (optional)",
                    "enum": ["North", "South", "East", "West", "All"]
                }
            }
        }
    
    async def generate(self, params: Dict[str, Any], output_path: str) -> str:
        days = params.get("days", 30)
        region_filter = params.get("region", "All")
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_font = Font(bold=True, size=14, color="2F5496")
        currency_format = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws_summary.merge_cells('A1:E1')
        title_cell = ws_summary['A1']
        title_cell.value = f"Sales Report - Last {days} days"
        title_cell.font = title_font
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A3'] = f"Region: {region_filter}"
        
        # Generate mock sales data
        products = ["Product A", "Product B", "Product C", "Product D", "Product E"]
        regions = ["North", "South", "East", "West"]
        if region_filter != "All":
            regions = [region_filter]
        
        # Summary metrics
        start_row = 5
        ws_summary[f'A{start_row}'] = "Key Metrics"
        ws_summary[f'A{start_row}'].font = title_font
        
        metrics = [
            ("Total Revenue", sum(random.uniform(10000, 50000) for _ in range(days * len(regions)))),
            ("Total Orders", random.randint(100, 500) * len(regions)),
            ("Average Order Value", random.uniform(100, 500)),
            ("Conversion Rate", random.uniform(2.0, 8.0)),
        ]
        
        for i, (metric, value) in enumerate(metrics, start=start_row + 1):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value
            if "Revenue" in metric or "Value" in metric:
                ws_summary[f'B{i}'].number_format = currency_format
        
        # Revenue by Product
        product_row = start_row + 7
        ws_summary[f'A{product_row}'] = "Revenue by Product"
        ws_summary[f'A{product_row}'].font = title_font
        
        headers = ["Product", "Units Sold", "Revenue", "Growth %"]
        header_start = product_row + 1
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=header_start, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for i, product in enumerate(products):
            row = header_start + 1 + i
            units = random.randint(50, 500)
            revenue = units * random.uniform(20, 200)
            growth = random.uniform(-10, 30)
            
            ws_summary.cell(row=row, column=1, value=product)
            ws_summary.cell(row=row, column=2, value=units)
            ws_summary.cell(row=row, column=3, value=revenue).number_format = currency_format
            ws_summary.cell(row=row, column=4, value=round(growth, 2))
            ws_summary.cell(row=row, column=4).number_format = '0.00"%"'
        
        # Revenue by Region
        region_row = product_row + len(products) + 3
        ws_summary[f'A{region_row}'] = "Revenue by Region"
        ws_summary[f'A{region_row}'].font = title_font
        
        for col, header in enumerate(["Region", "Orders", "Revenue", "% of Total"], 1):
            cell = ws_summary.cell(row=region_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        total_revenue = sum(random.uniform(10000, 50000) for _ in range(len(regions)))
        for i, region in enumerate(regions):
            row = region_row + 2 + i
            orders = random.randint(50, 300)
            revenue = random.uniform(5000, 25000)
            pct = (revenue / total_revenue) * 100
            
            ws_summary.cell(row=row, column=1, value=region)
            ws_summary.cell(row=row, column=2, value=orders)
            ws_summary.cell(row=row, column=3, value=revenue).number_format = currency_format
            ws_summary.cell(row=row, column=4, value=round(pct, 1))
            ws_summary.cell(row=row, column=4).number_format = '0.0"%"'
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 15
        
        # Daily Trends Sheet
        ws_trends = wb.create_sheet("Daily Trends")
        ws_trends['A1'] = "Date"
        ws_trends['B1'] = "Revenue"
        ws_trends['C1'] = "Orders"
        ws_trends['D1'] = "Avg Order Value"
        
        for col in range(1, 5):
            cell = ws_trends.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        base_date = datetime.now() - timedelta(days=days)
        for i in range(days):
            row = 2 + i
            date = base_date + timedelta(days=i)
            ws_trends.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
            ws_trends.cell(row=row, column=2, value=round(random.uniform(1000, 5000), 2)).number_format = currency_format
            ws_trends.cell(row=row, column=3, value=random.randint(10, 100))
            ws_trends.cell(row=row, column=4, value=round(random.uniform(50, 200), 2)).number_format = currency_format
        
        ws_trends.column_dimensions['A'].width = 15
        ws_trends.column_dimensions['B'].width = 15
        ws_trends.column_dimensions['C'].width = 12
        ws_trends.column_dimensions['D'].width = 18
        
        wb.save(output_path)
        return output_path
